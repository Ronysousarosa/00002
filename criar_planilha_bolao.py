
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def criar_planilha_bolao_lotofacil(nome_arquivo="bolao_lotofacil.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bolão Lotofácil"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Aba de Configurações
    ws_config = wb.create_sheet("Configurações")
    ws_config["A1"] = "Configurações do Bolão"
    ws_config["A1"].font = Font(bold=True, size=14)
    ws_config["A3"] = "Valor por Palpite (R$)"
    ws_config["B3"] = 5.00
    ws_config["B3"].number_format = '"R$"#,##0.00'
    ws_config["A4"] = "Senha de Administrador"
    ws_config["B4"] = "123456" # Apenas para referência, Excel não tem proteção de senha como o site
    ws_config.column_dimensions['A'].width = 25
    ws_config.column_dimensions['B'].width = 15

    # Aba de Resultados da Lotofácil
    ws_resultados = wb.create_sheet("Resultados Lotofácil")
    ws_resultados.append(["Concurso", "Data", "Dezenas Sorteadas"])
    for col in range(1, 4): # A, B, C
        ws_resultados.cell(row=1, column=col).font = header_font
        ws_resultados.cell(row=1, column=col).fill = header_fill
        ws_resultados.cell(row=1, column=col).alignment = center_align
    ws_resultados.column_dimensions['A'].width = 10
    ws_resultados.column_dimensions['B'].width = 12
    ws_resultados.column_dimensions['C'].width = 60
    
    # Exemplo de resultado
    ws_resultados.append([2500, "09/06/2025", "01, 02, 03, 04, 05, 06, 07, 08, 09, 10, 11, 12, 13, 14, 15"])
    
    # Aba Principal - Bolão Lotofácil
    headers = ["ID", "Nome do Participante", "Telefone", "Bilhete", "Dezenas Escolhidas", "Pontos", "Valor Pago (R$)", "Status Pagamento"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"].font = header_font
        ws[f"{col_letter}1"].fill = header_fill
        ws[f"{col_letter}1"].alignment = center_align
        ws.column_dimensions[col_letter].width = 20

    # Ajustes de largura para colunas específicas
    ws.column_dimensions['B'].width = 30 # Nome
    ws.column_dimensions['E'].width = 60 # Dezenas Escolhidas
    ws.column_dimensions['F'].width = 10 # Pontos
    ws.column_dimensions['G'].width = 15 # Valor Pago
    ws.column_dimensions['H'].width = 18 # Status Pagamento

    # Adicionar algumas linhas de exemplo
    ws.append([1, "João Silva", "(11) 98765-4321", "LF001", "01, 03, 05, 07, 09, 11, 13, 15, 17, 19", "=CONT.SE(Resultados_Lotofacil!C2:C2,A2)", "=Configurações!B3", "PAGO"])
    ws.append([2, "Maria Oliveira", "(21) 91234-5678", "LF002", "02, 04, 06, 08, 10, 12, 14, 16, 18, 20", "=CONT.SE(Resultados_Lotofacil!C2:C2,A3)", "=Configurações!B3", "PENDENTE"])
    
    # Adicionar fórmulas para cálculo de pontos (exemplo simplificado)
    # A fórmula real de contagem de pontos no Excel seria mais complexa
    # e exigiria uma função VBA ou uma fórmula matricial que não é trivial para gerar aqui.
    # A fórmula abaixo é um placeholder.
    # Para uma funcionalidade completa, o usuário precisaria de conhecimento em Excel VBA.
    
    # Aba de Resumo
    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo["A1"] = "Resumo do Bolão"
    ws_resumo["A1"].font = Font(bold=True, size=14)
    ws_resumo["A3"] = "Total de Participantes"
    ws_resumo["B3"] = "=CONT.VALORES('Bolão Lotofácil'!B:B)-1" # Exclui o cabeçalho
    ws_resumo["A4"] = "Valor Total Arrecadado (R$)"
    ws_resumo["B4"] = "=SOMA('Bolão Lotofácil'!G:G)"
    ws_resumo["B4"].number_format = '"R$"#,##0.00'
    ws_resumo["A5"] = "Maior Pontuação"
    ws_resumo["B5"] = "=MÁXIMO('Bolão Lotofácil'!F:F)"
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 20

    # Salvar a planilha
    wb.save(nome_arquivo)
    print(f"Planilha '{nome_arquivo}' criada com sucesso!")

if __name__ == "__main__":
    criar_planilha_bolao_lotofacil()

